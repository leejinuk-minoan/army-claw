from pathlib import Path

from openpyxl import Workbook, load_workbook

from openclaw.workspace import WorkspaceService
from openclaw.xlsx_tools import XlsxService


def create_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["부서", "금액"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])
    sheet.append(["A", 5])
    workbook.save(path)


def test_summarize_workbook(tmp_path: Path):
    create_sample_workbook(tmp_path / "sample.xlsx")
    service = XlsxService(WorkspaceService(tmp_path))

    summary = service.summarize_workbook("sample.xlsx")

    assert summary.sheets[0].name == "Data"
    assert summary.sheets[0].max_row == 4


def test_preview_sheet(tmp_path: Path):
    create_sample_workbook(tmp_path / "sample.xlsx")
    service = XlsxService(WorkspaceService(tmp_path))

    preview = service.preview_sheet("sample.xlsx", "Data")

    assert preview.rows[0] == ["부서", "금액"]
    assert preview.rows[1] == ["A", 10]


def test_write_cell(tmp_path: Path):
    create_sample_workbook(tmp_path / "sample.xlsx")
    service = XlsxService(WorkspaceService(tmp_path))

    result = service.write_cell("sample.xlsx", "Data", "C1", "비고")

    assert result.saved is True
    workbook = load_workbook(tmp_path / "sample.xlsx")
    assert workbook["Data"]["C1"].value == "비고"


def test_suggest_formula():
    service = XlsxService(WorkspaceService(Path.cwd()))

    suggestion = service.suggest_formula("sum", "B2:B4")

    assert suggestion.formula == "=SUM(B2:B4)"


def test_pivot_summary(tmp_path: Path):
    create_sample_workbook(tmp_path / "sample.xlsx")
    service = XlsxService(WorkspaceService(tmp_path))

    summary = service.pivot_summary("sample.xlsx", "Data", "부서", "금액")

    assert summary.values == {"A": 15.0, "B": 20.0}


def test_add_bar_chart(tmp_path: Path):
    create_sample_workbook(tmp_path / "sample.xlsx")
    service = XlsxService(WorkspaceService(tmp_path))

    result = service.add_bar_chart("sample.xlsx", "Data", "B1:B4")

    assert result.saved is True
