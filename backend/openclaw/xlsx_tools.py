from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from pydantic import BaseModel

from openclaw.workspace import WorkspaceError, WorkspaceService


class SheetSummary(BaseModel):
    name: str
    max_row: int
    max_column: int


class WorkbookSummary(BaseModel):
    path: str
    sheets: list[SheetSummary]


class SheetPreview(BaseModel):
    path: str
    sheet: str
    rows: list[list[Any]]


class CellWriteResult(BaseModel):
    path: str
    sheet: str
    cell: str
    value: Any
    saved: bool


class FormulaSuggestion(BaseModel):
    formula: str
    description: str


class PivotSummary(BaseModel):
    group_by: str
    value_column: str
    values: dict[str, float]


class ChartResult(BaseModel):
    path: str
    sheet: str
    chart_cell: str
    saved: bool


class XlsxService:
    def __init__(self, workspace: WorkspaceService) -> None:
        self.workspace = workspace

    def _resolve_xlsx(self, relative_path: str) -> Path:
        path = self.workspace.resolve_inside(relative_path)
        if path.suffix.lower() != ".xlsx":
            raise WorkspaceError("only .xlsx files are supported")
        return path

    def summarize_workbook(self, relative_path: str) -> WorkbookSummary:
        path = self._resolve_xlsx(relative_path)
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            return WorkbookSummary(
                path=relative_path,
                sheets=[
                    SheetSummary(name=sheet.title, max_row=sheet.max_row, max_column=sheet.max_column)
                    for sheet in workbook.worksheets
                ],
            )
        finally:
            workbook.close()

    def preview_sheet(self, relative_path: str, sheet_name: str, limit: int = 20) -> SheetPreview:
        path = self._resolve_xlsx(relative_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name]
            rows = [
                list(row)
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, limit),
                    values_only=True,
                )
            ]
            return SheetPreview(path=relative_path, sheet=sheet_name, rows=rows)
        finally:
            workbook.close()

    def write_cell(self, relative_path: str, sheet_name: str, cell: str, value: Any) -> CellWriteResult:
        path = self._resolve_xlsx(relative_path)
        workbook = load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            sheet[cell] = value
            workbook.save(path)
            return CellWriteResult(path=relative_path, sheet=sheet_name, cell=cell, value=value, saved=True)
        finally:
            workbook.close()

    def suggest_formula(self, function_name: str, cell_range: str) -> FormulaSuggestion:
        normalized = function_name.strip().upper()
        allowed = {
            "SUM": "합계를 계산합니다.",
            "AVERAGE": "평균을 계산합니다.",
            "COUNT": "숫자 셀 개수를 계산합니다.",
            "MAX": "최댓값을 계산합니다.",
            "MIN": "최솟값을 계산합니다.",
        }
        if normalized not in allowed:
            raise WorkspaceError(f"unsupported formula function: {function_name}")
        return FormulaSuggestion(formula=f"={normalized}({cell_range})", description=allowed[normalized])

    def pivot_summary(
        self,
        relative_path: str,
        sheet_name: str,
        group_by_column: str,
        value_column: str,
    ) -> PivotSummary:
        preview = self.preview_sheet(relative_path, sheet_name, limit=10_000)
        if not preview.rows:
            return PivotSummary(group_by=group_by_column, value_column=value_column, values={})

        headers = [str(value) if value is not None else "" for value in preview.rows[0]]
        try:
            group_index = headers.index(group_by_column)
            value_index = headers.index(value_column)
        except ValueError as exc:
            raise WorkspaceError("pivot columns must exist in the header row") from exc

        totals: dict[str, float] = defaultdict(float)
        for row in preview.rows[1:]:
            if group_index >= len(row) or value_index >= len(row):
                continue
            group_value = str(row[group_index])
            numeric_value = row[value_index]
            if isinstance(numeric_value, int | float):
                totals[group_value] += float(numeric_value)

        return PivotSummary(
            group_by=group_by_column,
            value_column=value_column,
            values=dict(sorted(totals.items())),
        )

    def add_bar_chart(
        self,
        relative_path: str,
        sheet_name: str,
        data_range: str,
        chart_cell: str = "H2",
        title: str = "Army Claw Chart",
    ) -> ChartResult:
        path = self._resolve_xlsx(relative_path)
        workbook = load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            min_col, min_row, max_col, max_row = self._parse_range(data_range)
            data = Reference(sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            chart = BarChart()
            chart.title = title
            chart.add_data(data, titles_from_data=True)
            sheet.add_chart(chart, chart_cell)
            workbook.save(path)
            return ChartResult(path=relative_path, sheet=sheet_name, chart_cell=chart_cell, saved=True)
        finally:
            workbook.close()

    def _parse_range(self, cell_range: str) -> tuple[int, int, int, int]:
        from openpyxl.utils.cell import range_boundaries

        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except ValueError as exc:
            raise WorkspaceError(f"invalid cell range: {cell_range}") from exc
        return min_col, min_row, max_col, max_row
